import openpyxl, os

# download example spreadsheet from:
# http://autbor.com/example.xlsx

os.chdir('c:\\users\\mkhir\\documents')

workbook = openpyxl.load_workbook('example.xlsx')
print(workbook)

sheet = workbook.get_sheet_by_name('Sheet1')
print(sheet)

print(workbook.get_sheet_names())

cell = sheet['A1']
print(str(cell.value))

print(sheet['B1'].value)
print(sheet['C1'].value)

# ***** #
sheet.cell(row=1, column=2)
# same as :
sheet['B1']
# ***** #

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)